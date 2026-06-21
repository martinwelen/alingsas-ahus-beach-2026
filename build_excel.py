#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Bygger ett Excel-dokument med alla matcher för Alingsås HK:s sex lag
på Åhus Beach Handboll 2026 (gruppspel). Källa: cupmanager.net (lagsidorna).

Matchdatan ligger i matches_data.py (gemensam källa med iCal-genereringen)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import schedule as sch
from matches_data import teams, team_colors

# ---------------------------------------------------------------------------
# Stilar
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color="1F3864", size=16)
SUB_FONT = Font(italic=True, color="555555", size=10)
ALI_FILL = PatternFill("solid", fgColor="FFF2CC")   # markera Alingsås-laget
ALI_FONT = Font(bold=True, color="9C5700")
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
sat_fill = PatternFill("solid", fgColor="EDEDED")

HEADERS = ["Datum", "Dag", "Tid", "Bana", "Lag (Alingsås)", "Klass",
           "Grupp", "Hemma", "Borta", "Motståndare", "Hemma/Borta"]
WIDTHS = [11, 9, 7, 7, 16, 16, 9, 26, 26, 24, 13]


def fmt_datum(d):
    y, m, day = d.split("-")
    return f"{int(day)}/{int(m)}"


def style_header(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = center
        cell.border = border


def write_match_table(ws, start_row, rows, color_by_team=False):
    for i, h in enumerate(HEADERS, start=1):
        ws.cell(row=start_row, column=i, value=h)
    style_header(ws, start_row, len(HEADERS))

    r = start_row + 1
    for m in rows:
        vals = [fmt_datum(m["datum"]), m["dag"], m["tid"], m["bana"],
                m["lag"], m["klass"], m["grupp"], m["hemma"], m["borta"],
                m["mots"], m["hb"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            cell.alignment = center if c in (1, 2, 3, 4, 6, 7, 11) else left
        if m["dag"] == "Lördag":
            for c in range(1, len(HEADERS) + 1):
                if ws.cell(row=r, column=c).fill.fgColor.rgb in (None, "00000000"):
                    ws.cell(row=r, column=c).fill = sat_fill
        ali_col = 8 if m["hb"] == "Hemma" else 9
        ws.cell(row=r, column=ali_col).fill = ALI_FILL
        ws.cell(row=r, column=ali_col).font = ALI_FONT
        if color_by_team:
            lc = ws.cell(row=r, column=5)
            lc.fill = PatternFill("solid", fgColor=team_colors[m["lag"]])
            lc.font = Font(bold=True, color="FFFFFF")
            lc.alignment = center
        r += 1
    return r - 1


def set_widths(ws):
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    wb = Workbook()
    all_rows, _meta = sch.load_matches()
    groups = sch.by_team(all_rows)
    all_sorted = sorted(all_rows, key=lambda m: (m["start_ms"], str(m["bana"])))

    # Blad 1: Alla matcher (kronologiskt)
    ws = wb.active
    ws.title = "Alla matcher"
    ws.cell(row=1, column=1, value="Alingsås HK – Åhus Beach Handboll 2026").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Alla gruppspelsmatcher för de sex lagen (3× P15 + 3× F15), "
                  "sorterade i tidsordning. Spelplats anges som bana-nummer.").font = SUB_FONT
    ws.cell(row=3, column=1,
            value="Speldagar: fredag 17 juli & lördag 18 juli 2026. "
                  "Slutspel (A/B/C) spelas på lördag em – se bladet 'Info & slutspel'.").font = SUB_FONT
    last = write_match_table(ws, 5, all_sorted, color_by_team=True)
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:{get_column_letter(len(HEADERS))}{last}"
    set_widths(ws)

    # Per-lag-blad
    for t in teams:
        ws = wb.create_sheet(title=t["lag"])
        ws.cell(row=1, column=1, value=t["fullnamn"]).font = TITLE_FONT
        ws.cell(row=2, column=1, value=f"{t['klass']} • {t['grupp']}").font = SUB_FONT
        rows = sorted(groups.get(t["slug"], []), key=lambda m: m["start_ms"])
        write_match_table(ws, 4, rows)
        ws.freeze_panes = "A5"
        set_widths(ws)
        ws.sheet_properties.tabColor = team_colors[t["lag"]]

    # Infoblad
    ws = wb.create_sheet(title="Info & slutspel")
    ws.column_dimensions["A"].width = 100
    info_lines = [
        ("Alingsås HK – Åhus Beach Handboll 2026", TITLE_FONT),
        ("", None),
        ("Detta dokument innehåller samtliga GRUPPSPELSMATCHER för Alingsås HK:s sex lag:", Font(bold=True)),
        ("   • Pojkar 15 år Classic (födda 2011): Blå (Grupp 2), Orange (Grupp 4), Vit (Grupp 1)", None),
        ("   • Flickor 15 år Classic (födda 2011): Blå (Grupp 1), Gul (Grupp 3), Vit (Grupp 4)", None),
        ("", None),
        ("Totalt antal gruppspelsmatcher: 51 st (P15: 9+9+9, F15: 8+8+8).", None),
        ("Speldagar: fredag 17 juli och lördag 18 juli 2026.", None),
        ("", None),
        ("SÅ LÄSER DU SCHEMAT:", Font(bold=True)),
        ("   • 'Bana' = spelplats/plan på området. Alingsås-laget är gulmarkerat i varje match.", None),
        ("   • 'Hemma/Borta' visar om Alingsås står som hemma- eller bortalag i matchprotokollet.", None),
        ("   • Bladet 'Alla matcher' är sorterat i tidsordning – bra för att se vad som spelas när.", None),
        ("   • Ett blad per lag (färgade flikar) för att följa ett enskilt lag.", None),
        ("", None),
        ("SLUTSPEL (A/B/C-slutspel):", Font(bold=True)),
        ("   Efter gruppspelet följer slutspel på lördag eftermiddag/kväll (ca kl 15:30 och framåt).", None),
        ("   Motståndare och exakta tider bestäms av gruppspelsresultaten och fanns ännu inte fastställda", None),
        ("   när dokumentet skapades. Kolla lagets sida på cupmanager.net under lördagen för uppdaterat slutspelsschema.", None),
        ("", None),
        ("Källa: ahusbeachhandboll.cupmanager.net (lagsidorna), hämtat 2026-06-21.", SUB_FONT),
        ("Tider och banor kan ändras av arrangören – dubbelkolla på plats eller i appen/sajten.", SUB_FONT),
    ]
    for i, (text, font) in enumerate(info_lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font

    out = "/home/martin/dev/ahusbeach/Alingsas_HK_Ahus_Beach_2026.xlsx"
    wb.save(out)
    print("Sparade", out)
    print("Antal matcher totalt:", len(all_rows))


if __name__ == "__main__":
    main()
